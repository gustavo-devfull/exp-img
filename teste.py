import os
import io
import datetime
import hashlib
import hmac
from openpyxl import load_workbook
from PIL import Image as PILImage
import requests

# =======================
# CONFIGURAÇÕES
# =======================
AWS_ACCESS_KEY_ID = 'DO00U3TGARCUQ4BBXLUF'
AWS_SECRET_ACCESS_KEY = '2UOswaN5G4JUnfv8wk/QTlO3KQU+5qywlnmoG8ho6kM'
REGION = 'nyc3'
SERVICE = 's3'
BUCKET = 'moribr'
ENDPOINT = f'https://{BUCKET}.{REGION}.digitaloceanspaces.com'
CDN_BASE = f'https://{BUCKET}.{REGION}.cdn.digitaloceanspaces.com/base-fotos/'
OUTPUT_FOLDER = 'imagens_exportadas'
EXCEL_FILE = 'ORDER.xlsx'

# =======================
# FUNÇÃO DE ASSINATURA AWS V4
# =======================
def sign(key, msg):
    return hmac.new(key, msg.encode('utf-8'), hashlib.sha256).digest()

def get_signature_key(key, dateStamp, regionName, serviceName):
    kDate = sign(('AWS4' + key).encode('utf-8'), dateStamp)
    kRegion = sign(kDate, regionName)
    kService = sign(kRegion, serviceName)
    kSigning = sign(kService, 'aws4_request')
    return kSigning

def upload_file(file_path, object_key):
    method = 'PUT'
    service = SERVICE
    host = f'{BUCKET}.{REGION}.digitaloceanspaces.com'
    content_type = 'image/jpeg'
    t = datetime.datetime.utcnow()
    amz_date = t.strftime('%Y%m%dT%H%M%SZ')
    datestamp = t.strftime('%Y%m%d')
    canonical_uri = f'/{object_key}'
    canonical_querystring = ''
    canonical_headers = f'host:{host}\nx-amz-content-sha256:UNSIGNED-PAYLOAD\nx-amz-date:{amz_date}\n'
    signed_headers = 'host;x-amz-content-sha256;x-amz-date'
    payload_hash = 'UNSIGNED-PAYLOAD'
    canonical_request = '\n'.join([method, canonical_uri, canonical_querystring,
                                   canonical_headers, signed_headers, payload_hash])
    algorithm = 'AWS4-HMAC-SHA256'
    credential_scope = f'{datestamp}/{REGION}/{service}/aws4_request'
    string_to_sign = '\n'.join([algorithm, amz_date, credential_scope,
                                hashlib.sha256(canonical_request.encode('utf-8')).hexdigest()])
    signing_key = get_signature_key(AWS_SECRET_ACCESS_KEY, datestamp, REGION, service)
    signature = hmac.new(signing_key, string_to_sign.encode('utf-8'),
                         hashlib.sha256).hexdigest()
    authorization_header = (f'{algorithm} Credential={AWS_ACCESS_KEY_ID}/{credential_scope}, '
                            f'SignedHeaders={signed_headers}, Signature={signature}')
    headers = {
        'x-amz-content-sha256': payload_hash,
        'x-amz-date': amz_date,
        'Authorization': authorization_header,
        'Content-Type': content_type
    }
    url = f'{ENDPOINT}/{object_key}'
    with open(file_path, 'rb') as f:
        response = requests.put(url, data=f, headers=headers)
    if response.status_code == 200:
        print(f'✅ Upload bem-sucedido: {url}')
    else:
        print(f'❌ Falha no upload: {url} - Status: {response.status_code}, Resposta: {response.text}')

# =======================
# EXTRAÇÃO DE IMAGENS
# =======================
def export_images_and_upload(excel_file, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    wb = load_workbook(excel_file)
    ws = wb.active

    for img in ws._images:
        if hasattr(img.anchor, '_from'):
            pos = img.anchor._from
            row = pos.row + 1
            ref_cell = ws[f"A{row}"].value
            if ref_cell is None:
                print(f'❌ Sem REF na linha {row}, pulando.')
                continue
            # Salvar imagem local
            img_bytes = io.BytesIO(img._data())
            pil_image = PILImage.open(img_bytes).convert('RGB')
            filename = f"{ref_cell}.jpg"
            file_path = os.path.join(output_folder, filename)
            pil_image.save(file_path, format='JPEG', quality=95)
            print(f'✅ Imagem salva: {file_path}')
            # Upload
            object_key = f'base-fotos/{filename}'
            upload_file(file_path, object_key)

# =======================
# MAIN
# =======================
if __name__ == "__main__":
    export_images_and_upload(EXCEL_FILE, OUTPUT_FOLDER)
