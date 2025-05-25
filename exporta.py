import os
import io
from openpyxl import load_workbook
from PIL import Image as PILImage
import boto3

# =======================
# Configurações AWS (DO Spaces)
# =======================
AWS_ACCESS_KEY_ID = 'DO00U3TGARCUQ4BBXLUF'
AWS_SECRET_ACCESS_KEY = '2UOswaN5G4JUnfv8wk/QTlO3KQU+5qywlnmoG8ho6kM'
AWS_REGION = 'nyc3'
BUCKET_NAME = 'moribr'
ENDPOINT_URL = 'https://moribr.nyc3.digitaloceanspaces.com'

# =======================
# Função de upload
# =======================
def upload_to_spaces(file_path, object_name):
    s3 = boto3.client('s3',
                      region_name=AWS_REGION,
                      endpoint_url=ENDPOINT_URL,
                      aws_access_key_id=AWS_ACCESS_KEY_ID,
                      aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
    s3.upload_file(file_path, BUCKET_NAME, object_name, ExtraArgs={'ACL': 'public-read'})
    print(f"Imagem {object_name} enviada para o Spaces.")

# =======================
# Processamento da planilha
# =======================
def export_images_and_upload(excel_file, output_folder='imagens_exportadas'):
    os.makedirs(output_folder, exist_ok=True)

    wb = load_workbook(excel_file)
    ws = wb.active

    for img in ws._images:
        if hasattr(img.anchor, '_from'):
            pos = img.anchor._from
            row = pos.row + 1
            ref_cell = ws[f"A{row}"].value

            if ref_cell is None:
                print(f"Sem REF para imagem na linha {row}, pulando...")
                continue

            # Exportar imagem
            img_bytes = io.BytesIO(img._data())
            pil_image = PILImage.open(img_bytes).convert('RGB')

            filename = f"{ref_cell}.jpg"
            file_path = os.path.join(output_folder, filename)

            pil_image.save(file_path, format='JPEG', quality=95)
            print(f"Imagem salva: {file_path}")

            # Fazer upload
            upload_to_spaces(file_path, f'base-fotos/{filename}')

if __name__ == "__main__":
    excel_file = 'cotacao.xlsx'  # Nome da sua planilha
    export_images_and_upload(excel_file)
